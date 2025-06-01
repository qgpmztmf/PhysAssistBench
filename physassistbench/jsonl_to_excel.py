"""
physassistbench/jsonl_to_excel.py — Convert PhysAssistBench benchmark JSONL to a readable Excel file.

Each row = one turn. Columns show:
  - Entry ID, Scenario, Sequence Index, Turn, Task Type, Subtype, Tool Source, Workup Mode
  - Question (EN/ZH), Gold Tools, Tool Args, Tool Observations, Gold Answer

Usage:
    cd /path/to/PhysAssistBench

    # Convert a single scenario file
    uv run python physassistbench/jsonl_to_excel.py physassistbench/data/infection_management.jsonl

    # Convert all scenario files into one Excel workbook
    uv run python physassistbench/jsonl_to_excel.py physassistbench/data/*.jsonl --output physassistbench/data/benchmark.xlsx

    # English columns only
    uv run python physassistbench/jsonl_to_excel.py physassistbench/data/infection_management.jsonl --language en

    # Chinese columns only
    uv run python physassistbench/jsonl_to_excel.py physassistbench/data/infection_management.jsonl --language zh
"""

import argparse
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ─── Colour palette ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(size=10)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER_TOP  = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _gold_tools(turn_actions: list) -> str:
    names = [
        a["action"]["name"]
        for a in turn_actions
        if a["action"]["name"] not in ("prepare_to_answer",)
    ]
    return ", ".join(names) if names else "—"


def _gold_answer(turn_actions: list) -> str:
    for a in turn_actions:
        if a["action"]["name"] == "prepare_to_answer":
            obs = a.get("observation", "")
            return str(obs) if obs else "—"
    return "—"


def _tool_args(turn_actions: list) -> str:
    lines = []
    for a in turn_actions:
        name = a["action"]["name"]
        if name == "prepare_to_answer":
            continue
        args = a["action"].get("arguments", {})
        display = {k: v for k, v in args.items() if k not in ("subject_id", "session_id", "hadm_id")}
        arg_str = ", ".join(f"{k}={v}" for k, v in display.items())
        lines.append(f"{name}({arg_str})" if arg_str else name)
    return "\n".join(lines) if lines else "—"


def _tool_observations(turn_actions: list) -> str:
    lines = []
    for a in turn_actions:
        name = a["action"]["name"]
        if name == "prepare_to_answer":
            continue
        obs = a.get("observation", "")
        obs_str = json.dumps(obs, ensure_ascii=False) if isinstance(obs, (dict, list)) else str(obs)
        if len(obs_str) > 500:
            obs_str = obs_str[:500] + "…"
        if name == "ask_user_for_required_parameters":
            user_inp = str(a.get("user_input", "")).strip()
            block = f"[{name}]\nClarification: {obs_str}"
            if user_inp:
                block += f"\nClinician response: {user_inp}"
            lines.append(block)
        elif name in ("ask_patient",):
            block = f"[{name}]\n{obs_str}"
            lines.append(block)
        else:
            lines.append(f"[{name}]\n{obs_str}")
    return "\n\n".join(lines) if lines else "—"


def _last_assistant_per_turn(messages: list, tasks: list) -> list[str]:
    """
    Extract the LAST assistant message per turn from messages_zh (no SEP markers).
    Uses tasks (ZH questions) to locate each turn's start, then takes the last
    assistant message before the next turn starts — which is the final answer,
    not the intermediate clarification question.
    """
    if not isinstance(messages, list) or not tasks:
        return []
    # Find the message index where each turn's user question appears
    turn_starts = []
    for task in tasks:
        task_stripped = str(task).strip()
        for idx, m in enumerate(messages):
            if (isinstance(m, dict)
                    and m.get("role") == "user"
                    and str(m.get("content", "")).strip() == task_stripped):
                turn_starts.append(idx)
                break
    result = []
    for i, start in enumerate(turn_starts):
        end = turn_starts[i + 1] if i + 1 < len(turn_starts) else len(messages)
        asst = [
            m.get("content", "")
            for m in messages[start:end]
            if isinstance(m, dict) and m.get("role") == "assistant"
        ]
        result.append(asst[-1] if asst else "—")
    return result


def build_rows(entries: list, language: str) -> list[dict]:
    rows = []
    for entry in entries:
        entry_id    = entry.get("id", "?")
        scenario    = entry.get("clinical_scenario", entry.get("clinical_task_domain", "?"))
        seq_idx     = entry.get("sequence_idx", "?")
        task_seq    = entry.get("task_sequence", [])
        task_types  = entry.get("task_types", [])
        subtypes    = entry.get("turn_subtypes", [])
        tool_sources= entry.get("tool_sources", [])
        workup_modes= entry.get("workup_modes", [])

        tasks_en    = entry.get("tasks_en") or entry.get("tasks", [])
        tasks_zh    = entry.get("tasks_zh", [])
        answer_list = entry.get("answer_list", [])

        answers_zh_by_turn = _last_assistant_per_turn(
            entry.get("messages_zh", []), tasks_zh
        )

        n_turns = max(len(tasks_en), len(answer_list))

        for i in range(n_turns):
            task_type   = task_types[i]   if i < len(task_types)   else "?"
            subtype     = subtypes[i]     if i < len(subtypes)     else "—"
            tool_source = tool_sources[i] if i < len(tool_sources) else "?"
            workup_mode = workup_modes[i] if i < len(workup_modes) else "—"
            gold        = answer_list[i]  if i < len(answer_list)  else []

            row = {
                "Entry ID":     entry_id,
                "Scenario":     scenario,
                "Seq Idx":      seq_idx,
                "Turn":         i,
                "Task Type":    task_type,
                "Subtype":      subtype or "—",
                "Tool Source":  tool_source,
                "Workup Mode":  workup_mode or "—",
            }

            if language in ("en", "both"):
                row["Question (EN)"]         = tasks_en[i] if i < len(tasks_en) else "—"
                row["Gold Tools (EN)"]        = _gold_tools(gold)
                row["Tool Calls (EN)"]        = _tool_args(gold)
                row["Tool Observations (EN)"] = _tool_observations(gold)
                row["Answer (EN)"]            = _gold_answer(gold)

            if language in ("zh", "both"):
                row["Question (ZH)"]         = tasks_zh[i] if i < len(tasks_zh) else "—"
                row["Gold Tools (ZH)"]        = _gold_tools(gold)
                row["Tool Calls (ZH)"]        = _tool_args(gold)
                row["Tool Observations (ZH)"] = _tool_observations(gold)
                row["Answer (ZH)"]            = (
                    answers_zh_by_turn[i] if i < len(answers_zh_by_turn)
                    else _gold_answer(gold)
                )

            rows.append(row)
    return rows


def write_sheet(ws, rows: list[dict], sheet_title: str):
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_TOP

    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.font = NORMAL_FONT
            cell.fill = fill
            cell.alignment = WRAP

    col_widths = {
        "Entry ID": 32, "Scenario": 20, "Seq Idx": 8, "Turn": 6,
        "Task Type": 22, "Subtype": 8, "Tool Source": 10, "Workup Mode": 12,
        "Question (EN)": 45, "Gold Tools (EN)": 30, "Tool Calls (EN)": 35,
        "Tool Observations (EN)": 60, "Answer (EN)": 55,
        "Question (ZH)": 45, "Gold Tools (ZH)": 30, "Tool Calls (ZH)": 35,
        "Tool Observations (ZH)": 60, "Answer (ZH)": 55,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 18)

    ws.freeze_panes = "A2"


def convert(input_files: list[str], output_path: str, language: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for fpath in input_files:
        if not os.path.exists(fpath):
            print(f"  SKIP (not found): {fpath}")
            continue

        with open(fpath, encoding="utf-8") as f:
            entries = [json.loads(line) for line in f if line.strip()]

        sheet_name = os.path.splitext(os.path.basename(fpath))[0][:31]
        ws = wb.create_sheet(title=sheet_name)
        rows = build_rows(entries, language=language)
        write_sheet(ws, rows, sheet_title=sheet_name)
        print(f"  {sheet_name}: {len(entries)} entries, {len(rows)} turns → sheet '{sheet_name}'")

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert PhysAssistBench benchmark JSONL files to a readable Excel workbook."
    )
    parser.add_argument("inputs", nargs="+", help="One or more .jsonl files to convert")
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output .xlsx path (default: benchmark.xlsx in same dir as first input)",
    )
    parser.add_argument(
        "--language", default="both", choices=["en", "zh", "both"],
        help="Which language columns to include (default: both)",
    )
    args = parser.parse_args()

    if args.output is None:
        out_dir = os.path.dirname(os.path.abspath(args.inputs[0]))
        args.output = os.path.join(out_dir, "benchmark.xlsx")

    print(f"Language: {args.language}")
    convert(args.inputs, args.output, language=args.language)


if __name__ == "__main__":
    main()
