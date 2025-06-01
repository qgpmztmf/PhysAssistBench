"""
Export annotation review data to Excel.
10 entries per scenario, stratified by subtype coverage.
Each row = one turn.
"""
import json
import glob
import random
import os
import sys
random.seed(42)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SCENARIOS = ["diagnostic_workup", "med_safety", "treatment_response", "discharge_planning"]
DATA_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_full")
OUT_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "annotation_review.xlsx")
N_PER_SCENARIO = 10

SUBTYPE_COLORS = {
    None:  "FFFFFF",   # white  — explicit
    "PE":  "DBEAFE",   # blue
    "NA":  "DCFCE7",   # green
    "AE":  "FEF9C3",   # yellow
}

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
SUBHDR_FILL  = PatternFill("solid", fgColor="2D6A9F")
SUBHDR_FONT  = Font(bold=True, color="FFFFFF", size=9)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def sample_entries(entries, n=10):
    """Stratified sample: cover all 3 difficulty levels + max subtype variety."""
    by_diff = {1: [], 2: [], 3: []}
    for e in entries:
        by_diff[e.get("difficulty_level", 1)].append(e)

    sampled = []
    per_diff = max(1, n // 3)
    for d in [1, 2, 3]:
        pool = by_diff.get(d, [])
        random.shuffle(pool)
        sampled.extend(pool[:per_diff])

    # Fill remaining slots, prefer entries with NA or AE subtypes
    needed = n - len(sampled)
    sampled_ids = {e["id"] for e in sampled}
    remaining = [e for e in entries if e["id"] not in sampled_ids]
    # Prioritize NA/AE coverage
    has_rare = [e for e in remaining
                if any(s in ("NA", "AE") for s in e.get("turn_subtypes", []))]
    others   = [e for e in remaining if e not in has_rare]
    random.shuffle(has_rare); random.shuffle(others)
    sampled.extend((has_rare + others)[:needed])
    return sampled[:n]


def _get_gold_answer(entry: dict, ti: int) -> str:
    """Extract gold answer from prepare_to_answer observation."""
    answer_list = entry.get("answer_list", [])
    if ti >= len(answer_list):
        return ""
    for act in answer_list[ti]:
        if act.get("action", {}).get("name") == "prepare_to_answer":
            obs = act.get("observation", "")
            return str(obs) if obs else ""
    return ""


def _get_gold_answer_zh(entry: dict, ti: int) -> str:
    """Extract Chinese gold answer from messages_zh."""
    msgs_zh = entry.get("messages_zh", [])
    # Filter out separator strings (they're inserted between turns)
    dict_msgs = [m for m in msgs_zh if isinstance(m, dict)]
    asst_idx = ti * 2 + 1   # user, assistant, user, assistant, ...
    if asst_idx < len(dict_msgs):
        m = dict_msgs[asst_idx]
        if m.get("role") == "assistant":
            return m.get("content", "")
    return ""


def build_rows(entries):
    rows = []
    for e in entries:
        subtypes   = e.get("turn_subtypes", [])
        tasks_en   = e.get("tasks_en") or []
        tasks_zh   = e.get("tasks_zh") or []
        tasks_en_x = e.get("tasks_en_explicit") or []
        tasks_zh_x = e.get("tasks_zh_explicit") or []
        rubrics_en = e.get("rubrics") or []
        rubrics_zh = e.get("rubrics_zh") or []
        task_types = e.get("task_types") or []

        for ti in range(len(tasks_en)):
            st      = subtypes[ti] if ti < len(subtypes) else None
            tt      = task_types[ti] if ti < len(task_types) else ""
            im_en   = tasks_en[ti]   if ti < len(tasks_en)   else ""
            im_zh   = tasks_zh[ti]   if ti < len(tasks_zh)   else ""
            ex_en   = tasks_en_x[ti] if ti < len(tasks_en_x) else ""
            ex_zh   = tasks_zh_x[ti] if ti < len(tasks_zh_x) else ""
            r_en    = rubrics_en[ti] if ti < len(rubrics_en)  else []
            r_zh    = rubrics_zh[ti] if ti < len(rubrics_zh)  else []
            answer_zh = _get_gold_answer_zh(e, ti) or _get_gold_answer(e, ti)

            rubric_en_str = "\n".join(f"{i+1}. {item}" for i, item in enumerate(r_en))
            rubric_zh_str = "\n".join(f"{i+1}. {item}" for i, item in enumerate(r_zh))

            rows.append({
                "entry_id":    e["id"],
                "scenario":    e.get("clinical_scenario", ""),
                "difficulty":  e.get("difficulty_level", ""),
                "turn":        ti,
                "task_type":   tt,
                "explicit_en": ex_en,
                "implicit_en": im_en,
                "rubric_en":   rubric_en_str,
                "rubric_zh":   rubric_zh_str,
                # last 4 columns (answer = Chinese)
                "subtype":     st if st is not None else "Explicit",
                "explicit_zh": ex_zh,
                "implicit_zh": im_zh,
                "answer":      answer_zh,
                # QC review columns (empty — to be filled by annotators)
                "need_modify":   "",
                "mod_explicit_zh": "",
                "mod_implicit_zh": "",
                "_st":         st,
                "_last_in_entry": (ti == len(tasks_en) - 1),
            })
    return rows


def write_excel(all_rows, out_path):
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Scenarios"

    scenario_sheets = {}
    for sc in SCENARIOS:
        ws = wb.create_sheet(title=sc[:20])
        scenario_sheets[sc] = ws

    COLUMNS = [
        ("Entry ID",      "entry_id",    28),
        ("Scenario",      "scenario",    18),
        ("Difficulty",    "difficulty",   9),
        ("Turn",          "turn",         6),
        ("Task Type",     "task_type",   18),
        ("Explicit EN",   "explicit_en", 45),
        ("Implicit EN",   "implicit_en", 45),
        ("Rubric EN",     "rubric_en",   60),
        ("Rubric ZH",     "rubric_zh",   60),
        # ZH content
        ("Subtype",       "subtype",     10),
        ("Explicit ZH",   "explicit_zh", 45),
        ("Implicit ZH",   "implicit_zh", 45),
        ("Answer (ZH)",   "answer",      60),
        # QC review columns
        ("Need to modify",      "need_modify",     14),
        ("Modified Explicit ZH", "mod_explicit_zh", 45),
        ("Modified Implicit ZH", "mod_implicit_zh", 45),
    ]

    def write_header(ws):
        # Row 1: main header
        ws.append([c[0] for c in COLUMNS])
        for col_idx, (_, _, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def write_row(ws, row_num, row_data):
        st = row_data["_st"]
        bg = SUBTYPE_COLORS.get(st, "FFFFFF")
        fill = PatternFill("solid", fgColor=bg)
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _thin_border()
            cell.font = Font(size=9)

    def write_rows_with_separator(ws, rows):
        """Write rows with an empty separator row after each 4-turn entry."""
        row_num = 2
        for row in rows:
            write_row(ws, row_num, row)
            row_num += 1
            if row.get("_last_in_entry"):
                # Empty separator row — light grey background
                sep_fill = PatternFill("solid", fgColor="F3F4F6")
                for col_idx in range(1, len(COLUMNS) + 1):
                    cell = ws.cell(row=row_num, column=col_idx, value="")
                    cell.fill = sep_fill
                ws.row_dimensions[row_num].height = 18
                row_num += 1

    # Write All sheet
    write_header(ws_all)
    write_rows_with_separator(ws_all, all_rows)

    # Write per-scenario sheets
    for sc, ws in scenario_sheets.items():
        write_header(ws)
        sc_rows = [r for r in all_rows if r["scenario"] == sc]
        write_rows_with_separator(ws, sc_rows)

    # Legend sheet
    ws_leg = wb.create_sheet("Legend")
    ws_leg.column_dimensions["A"].width = 15
    ws_leg.column_dimensions["B"].width = 55
    legend_data = [
        ("Color", "Meaning"),
        ("White",  "Explicit (Subtype=None) — T0 or Write/Update turns; no implicit reference"),
        ("Blue",   "PE: Predicate Ellipsis — verb phrase omitted (e.g. 'Latest creatinine?')"),
        ("Green",  "NA: Nominal Anaphora — noun replaced by pronoun (e.g. 'With it...')"),
        ("Yellow", "AE: Abstract Event Anaphora — refers to prior clinical event ('Given all that...')"),
    ]
    for ri, (k, v) in enumerate(legend_data, 1):
        ws_leg.cell(ri, 1, k).font = Font(bold=(ri==1), size=10)
        ws_leg.cell(ri, 2, v).font = Font(size=10)
        colors = {2: "FFFFFF", 3: "DBEAFE", 4: "DCFCE7", 5: "FEF9C3"}
        if ri in colors:
            ws_leg.cell(ri, 1).fill = PatternFill("solid", fgColor=colors[ri])
            ws_leg.cell(ri, 2).fill = PatternFill("solid", fgColor=colors[ri])

    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--all", action="store_true",
                        help="Export all entries (no sampling)")
    parser.add_argument("--out", default=None,
                        help="Output path override")
    args = parser.parse_args()

    out_path = args.out or (
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "annotation_review_all.xlsx")
        if args.all else OUT_PATH
    )

    all_rows = []
    from collections import Counter

    for sc in SCENARIOS:
        path = os.path.join(DATA_DIR, f"{sc}.jsonl")
        with open(path) as f:
            entries = [json.loads(l) for l in f if l.strip()]
        if args.all:
            selected = sorted(entries, key=lambda e: (e.get("difficulty_level", 0),
                                                       e.get("id", "")))
        else:
            selected = sample_entries(entries, N_PER_SCENARIO)
        rows = build_rows(selected)
        all_rows.extend(rows)
        st_cnt = Counter(r["subtype"] for r in rows)
        print(f"  {sc}: {len(selected)} entries, {len(rows)} turns  {dict(st_cnt)}")

    print(f"\nTotal rows: {len(all_rows)}")
    write_excel(all_rows, out_path)


if __name__ == "__main__":
    main()
