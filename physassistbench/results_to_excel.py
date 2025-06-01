"""
physassistbench/results_to_excel.py — Export eval results to a detailed Excel file.

Each row = one turn. Includes:
  - Entry metadata (ID, scenario, domain, sequence index)
  - Turn detail (turn #, task type, subtype, tool source)
  - User question + LLM answer + gold answer
  - LLM tool calls vs gold tool calls
  - Tool metrics (AP rate, extra/missed tools)
  - IIRS components (AID, ER, AC, AA, IIRS_turn)
  - is_correct, is_correct_method, is_correct_reason
  - Session accuracy (all 4 turns correct?)

Usage:
    cd /path/to/PhysAssistBench
    uv run python physassistbench/results_to_excel.py
    uv run python physassistbench/results_to_excel.py --output physassistbench/results_en/my_results.xlsx
"""

import argparse
import json
import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="2F5496")
CORRECT_FILL  = PatternFill("solid", fgColor="C6EFCE")   # green
WRONG_FILL    = PatternFill("solid", fgColor="FFC7CE")    # red
SESSION_FILL  = PatternFill("solid", fgColor="FFEB9C")    # yellow — session correct
ALT_FILL      = PatternFill("solid", fgColor="DCE6F1")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT   = Font(size=10)
BOLD_FONT     = Font(bold=True, size=10)
WRAP          = Alignment(wrap_text=True, vertical="top")
CENTER_TOP    = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _tool_names(actions: list) -> str:
    names = [
        a["action"]["name"] for a in actions
        if a["action"]["name"] != "prepare_to_answer"
    ]
    return ", ".join(names) if names else "—"


def _tool_args(actions: list) -> str:
    lines = []
    for a in actions:
        name = a["action"]["name"]
        if name == "prepare_to_answer":
            continue
        args = a["action"].get("arguments", {})
        display = {k: v for k, v in args.items()
                   if k not in ("subject_id", "hadm_id", "session_id")}
        arg_str = ", ".join(f"{k}={v}" for k, v in display.items())
        lines.append(f"{name}({arg_str})" if arg_str else name)
    return "\n".join(lines) if lines else "—"


def _gold_answer(actions: list) -> str:
    for a in actions:
        if a["action"]["name"] == "prepare_to_answer":
            obs = a.get("observation", "")
            return str(obs) if obs else "—"
    return "—"


def _truncate(text: str, n: int = 800) -> str:
    s = str(text)
    return s[:n] + "…" if len(s) > n else s


def build_rows(turn_results: list, session_correct_map: dict) -> list[dict]:
    rows = []
    for r in turn_results:
        entry_id   = r.get("test_entry_id", "?")
        # Parse scenario from entry_id: PhysAssistBench_Domain_scenario_N
        parts = entry_id.split("_")
        scenario = "_".join(parts[2:-1]) if len(parts) >= 4 else entry_id

        llm_actions  = r.get("llm_executed_actions") or r.get("executed_actions", [])
        gold_actions = r.get("gold_actions", [])

        rows.append({
            "Entry ID":            entry_id,
            "Scenario":            scenario,
            "Domain":              r.get("clinical_task_domain", "?"),
            "Turn":                r.get("task_idx", "?"),
            "Task Type":           r.get("task_type", "?"),
            "Subtype":             r.get("turn_subtype") or "None",
            # Questions & Answers
            "User Question":       r.get("user_question", ""),
            "LLM Answer":          _truncate(r.get("llm_answer", "") or ""),
            "Gold Answer":         _truncate(_gold_answer(gold_actions)),
            # Tool calls
            "LLM Tools":           _tool_names(llm_actions),
            "LLM Tool Args":       _tool_args(llm_actions),
            "Gold Tools":          _tool_names(gold_actions),
            "Gold Tool Args":      _tool_args(gold_actions),
            # Tool metrics
            "AP Rate":             round(r.get("ap_rate", 0), 3),
            "Tool Names Correct":  r.get("tool_names_correct", False),
            "Tool Coverage":       r.get("tool_coverage_correct", False),
            "Extra Tools":         ", ".join(r.get("extra_tools", [])) or "—",
            "Missed Tools":        ", ".join(r.get("missed_tools", [])) or "—",
            # IIRS
            "AID":                 round(r.get("iirs_AID", 0), 3),
            "ER":                  round(r.get("iirs_ER", 0), 3),
            "AC":                  round(r.get("iirs_AC", 0), 3),
            "AA":                  round(r.get("iirs_AA", 0), 3),
            "IIRS Turn":           round(r.get("iirs_turn", 0), 3),
            "FPTR":                round(r.get("fptr", 0), 3),
            "MTR":                 round(r.get("mtr", 0), 3),
            # Correctness
            "Is Correct":          r.get("is_correct", None),
            "Correct Method":      r.get("is_correct_method", "—"),
            "Correct Reason":      _truncate(r.get("is_correct_reason", ""), 200),
            # Session
            "Session Correct":     session_correct_map.get(entry_id, None),
        })
    return rows


def write_sheet(ws, rows: list[dict]):
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_TOP

    prev_entry = None
    alt = False
    for row_idx, row in enumerate(rows, 2):
        entry_id = row["Entry ID"]
        if entry_id != prev_entry:
            alt = not alt
            prev_entry = entry_id

        is_correct  = row.get("Is Correct")
        sess_correct = row.get("Session Correct")

        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.font = NORMAL_FONT
            cell.alignment = WRAP

            # Row background: correct=green, wrong=red, else alternating
            if is_correct is True:
                cell.fill = CORRECT_FILL
            elif is_correct is False:
                cell.fill = WRONG_FILL
            else:
                cell.fill = ALT_FILL if alt else WHITE_FILL

    col_widths = {
        "Entry ID": 36, "Scenario": 22, "Domain": 14,
        "Turn": 6, "Task Type": 22, "Subtype": 8,
        "User Question": 45, "LLM Answer": 55, "Gold Answer": 55,
        "LLM Tools": 30, "LLM Tool Args": 35,
        "Gold Tools": 30, "Gold Tool Args": 35,
        "AP Rate": 9, "Tool Names Correct": 14, "Tool Coverage": 12,
        "Extra Tools": 25, "Missed Tools": 25,
        "AID": 7, "ER": 7, "AC": 7, "AA": 7, "IIRS Turn": 10,
        "FPTR": 7, "MTR": 7,
        "Is Correct": 10, "Correct Method": 12, "Correct Reason": 40,
        "Session Correct": 14,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)

    ws.freeze_panes = "A2"


def write_summary_sheet(ws, turn_results: list, session_correct_map: dict):
    """Aggregate stats sheet."""
    total = len(turn_results)
    n_correct = sum(1 for r in turn_results if r.get("is_correct"))
    n_sessions = len(session_correct_map)
    n_sess_correct = sum(1 for v in session_correct_map.values() if v)

    by_type: dict = defaultdict(list)
    by_subtype: dict = defaultdict(list)
    by_scenario: dict = defaultdict(list)
    by_turn: dict = defaultdict(list)

    for r in turn_results:
        by_type[r.get("task_type", "?")].append(r)
        by_subtype[r.get("turn_subtype") or "None"].append(r)
        parts = r.get("test_entry_id", "").split("_")
        scenario = "_".join(parts[2:-1]) if len(parts) >= 4 else "?"
        by_scenario[scenario].append(r)
        by_turn[r.get("task_idx", 0)].append(r)

    def acc(items):
        if not items:
            return "—"
        n = sum(1 for r in items if r.get("is_correct"))
        return f"{n}/{len(items)} ({100*n/len(items):.1f}%)"

    def iirs_mean(items):
        vals = [r.get("iirs_turn", 0) for r in items if r.get("iirs_turn") is not None]
        return f"{sum(vals)/len(vals):.3f}" if vals else "—"

    rows = [
        ["Metric", "Value"],
        ["Total turns", total],
        ["Turn accuracy", acc(turn_results)],
        ["Sessions", n_sessions],
        ["Session accuracy", f"{n_sess_correct}/{n_sessions} ({100*n_sess_correct/n_sessions:.1f}%)" if n_sessions else "—"],
        ["IIRS (mean)", iirs_mean(turn_results)],
        [],
        ["By Task Type", "Turn Accuracy", "IIRS (mean)"],
    ]
    for tt in ["Information Lookup", "Data Gathering", "Clinical Reasoning"]:
        items = by_type.get(tt, [])
        rows.append([tt, acc(items), iirs_mean(items)])

    rows += [[], ["By Subtype", "Turn Accuracy", "IIRS (mean)"]]
    for st in ["None", "EA", "PE", "AU", "AE"]:
        items = by_subtype.get(st, [])
        if items:
            rows.append([st, acc(items), iirs_mean(items)])

    rows += [[], ["By Turn Index", "Turn Accuracy", "IIRS (mean)"]]
    for idx in sorted(by_turn.keys()):
        rows.append([f"Turn {idx}", acc(by_turn[idx]), iirs_mean(by_turn[idx])])

    rows += [[], ["By Scenario", "Turn Accuracy", "IIRS (mean)"]]
    for sc in sorted(by_scenario.keys()):
        rows.append([sc, acc(by_scenario[sc]), iirs_mean(by_scenario[sc])])

    for row_idx, row in enumerate(rows, 1):
        ws.append(row)
        if row and row[0] in ("Metric", "By Task Type", "By Subtype", "By Turn Index", "By Scenario"):
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill("solid", fgColor="2F5496")
                cell.font = HEADER_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15


# ── Subtype descriptions for the examples sheet ───────────────────────────────
_SUBTYPE_DESC = {
    "None": "Turn 0 — explicit question with no prior context to resolve",
    "EA":   "Entity Anaphora — a pronoun/demonstrative refers to a specific entity named in a prior turn (e.g. 'What about it?')",
    "PE":   "Predicate Ellipsis — the action is omitted; only the target is stated (e.g. 'And the albumin?')",
    "AU":   "Argument Underspecification — the action is vague; arguments must be inferred from context (e.g. 'Get the latest labs')",
    "AE":   "Abstract/Event Anaphora — refers to an abstract clinical situation built across multiple turns (e.g. 'Given all that...')",
}

_TASK_DESC = {
    "Information Lookup":           "Call exactly 1 EHR or patient tool to retrieve a specific data point",
    "Data Gathering":              "Call 2+ tools (parallel or conditional) to gather a comprehensive clinical picture",
    "Clinical Reasoning":  "Retrieve 1 patient parameter, then apply clinical knowledge to give personalised advice",
}

_APPLIES = {
    "None": "AC + AA",
    "EA":   "AID + AA",
    "PE":   "ER + AA",
    "AU":   "ER + AC + AA",
    "AE":   "AID + AA",
}


def write_examples_sheet(ws, turn_results: list):
    """
    One full 4-turn session example per task-type × subtype combination.
    The session is chosen as the one whose target turn has the highest IIRS score.
    All 4 turns of that session are shown, with the target turn highlighted.
    """
    # Build lookups
    by_entry: dict = defaultdict(list)   # entry_id → [turn0, turn1, turn2, turn3]
    for r in turn_results:
        by_entry[r.get("test_entry_id", "")].append(r)
    for turns in by_entry.values():
        turns.sort(key=lambda x: x.get("task_idx", 0))

    # Group by (task_type, subtype)
    combos: dict = defaultdict(list)
    for r in turn_results:
        key = (r.get("task_type", "?"), r.get("turn_subtype") or "None")
        combos[key].append(r)

    TASK_ORDER    = ["Information Lookup", "Data Gathering", "Clinical Reasoning"]
    SUBTYPE_ORDER = ["None", "EA", "PE", "AU", "AE"]

    # Turn background colours (one per turn index)
    TURN_FILLS = ["FFF2CC", "DEEBF7", "E2EFDA", "FCE4D6"]  # yellow/blue/green/orange

    row_num = 1

    def label_cell(row, col, label, value, label_fill="E2EFDA", n_merge=7):
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = Font(bold=True, size=10)
        lc.fill = PatternFill("solid", fgColor=label_fill)
        lc.alignment = WRAP
        vc = ws.cell(row=row, column=col + 1, value=value)
        vc.font = NORMAL_FONT
        vc.alignment = WRAP
        if n_merge > 1:
            ws.merge_cells(start_row=row, start_column=col+1,
                           end_row=row, end_column=col+n_merge)

    # Title
    tc = ws.cell(row=row_num, column=1,
                 value="Examples: Full 4-Turn Session per Task Type × Subtype Combination")
    tc.fill = PatternFill("solid", fgColor="1F3864")
    tc.font = Font(bold=True, size=14, color="FFFFFF")
    tc.alignment = WRAP
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
    row_num += 2

    for task_type in TASK_ORDER:
        for subtype in SUBTYPE_ORDER:
            key = (task_type, subtype)
            candidates = combos.get(key, [])
            if not candidates:
                continue

            # Pick the turn with highest IIRS, then use its full session
            target_turn = max(candidates, key=lambda x: x.get("iirs_turn", 0) or 0)
            entry_id    = target_turn.get("test_entry_id", "")
            target_idx  = target_turn.get("task_idx", 0)
            session_turns = by_entry.get(entry_id, [target_turn])

            # ── Section header ─────────────────────────────────────────────
            header_text = f"{task_type}  /  Subtype: {subtype}  (session: {entry_id})"
            hc = ws.cell(row=row_num, column=1, value=header_text)
            hc.font = Font(bold=True, size=12, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor="2F5496")
            hc.alignment = WRAP
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            row_num += 1

            # ── Concept block ──────────────────────────────────────────────
            label_cell(row_num, 1, "Task Type",   _TASK_DESC.get(task_type, ""), "E2EFDA")
            row_num += 1
            label_cell(row_num, 1, "Subtype",     _SUBTYPE_DESC.get(subtype, ""), "E2EFDA")
            row_num += 1
            label_cell(row_num, 1, "IIRS applies", _APPLIES.get(subtype, "AC + AA"), "E2EFDA")
            row_num += 1
            row_num += 1  # spacer

            # ── All 4 turns ────────────────────────────────────────────────
            for t in session_turns:
                tidx         = t.get("task_idx", 0)
                tt           = t.get("task_type", "?")
                st           = t.get("turn_subtype") or "None"
                is_target    = (tidx == target_idx)
                turn_fill    = TURN_FILLS[tidx % len(TURN_FILLS)]
                highlight    = "FFD700" if is_target else turn_fill   # gold highlight for target

                llm_actions  = t.get("llm_executed_actions") or t.get("executed_actions", [])
                gold_actions = t.get("gold_actions", [])

                # Turn sub-header
                marker = "  ◀ THIS TURN is the example for this task/subtype" if is_target else ""
                turn_label = f"Turn {tidx}  [{tt} / {st}]{marker}"
                thc = ws.cell(row=row_num, column=1, value=turn_label)
                thc.font = Font(bold=True, size=11, color="FFFFFF" if is_target else "000000")
                thc.fill = PatternFill("solid", fgColor="FF6600" if is_target else "70AD47")
                thc.alignment = WRAP
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
                row_num += 1

                label_cell(row_num, 1, "Question",     t.get("user_question", ""), highlight)
                row_num += 1
                label_cell(row_num, 1, "Gold Tools",   _tool_names(gold_actions), highlight)
                row_num += 1
                label_cell(row_num, 1, "Gold Args",    _tool_args(gold_actions), highlight)
                row_num += 1
                label_cell(row_num, 1, "Gold Answer",  _truncate(_gold_answer(gold_actions), 400), highlight)
                row_num += 1
                label_cell(row_num, 1, "LLM Tools",    _tool_names(llm_actions), "DDEBF7")
                row_num += 1
                label_cell(row_num, 1, "LLM Args",     _tool_args(llm_actions), "DDEBF7")
                row_num += 1
                label_cell(row_num, 1, "LLM Answer",   _truncate(t.get("llm_answer", ""), 400), "DDEBF7")
                row_num += 1

                scores_text = (
                    f"AID={t.get('iirs_AID',0):.2f}  "
                    f"ER={t.get('iirs_ER',0):.2f}  "
                    f"AC={t.get('iirs_AC',0):.2f}  "
                    f"AA={t.get('iirs_AA',0):.2f}  "
                    f"IIRS={t.get('iirs_turn',0):.3f}  "
                    f"Is Correct={'✓' if t.get('is_correct') else '✗'}"
                )
                label_cell(row_num, 1, "Scores", scores_text, "F2F2F2")
                row_num += 1
                if t.get("is_correct_reason"):
                    label_cell(row_num, 1, "Reason",
                               _truncate(t.get("is_correct_reason", ""), 200), "F2F2F2")
                    row_num += 1
                row_num += 1  # gap between turns

            row_num += 2  # gap between sections

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 95
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 0


def main():
    parser = argparse.ArgumentParser(description="Export PhysAssistBench eval results to Excel")
    parser.add_argument(
        "--results_dir", default=None,
        help="Path to results directory (default: physassistbench/results_en/)",
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output .xlsx path (default: results_en/eval_results.xlsx)",
    )
    args = parser.parse_args()

    base = os.path.dirname(os.path.abspath(__file__))
    results_dir = args.results_dir or os.path.join(base, "results_en")
    final_path = os.path.join(results_dir, "turn_results_final.json")

    if not os.path.exists(final_path):
        print(f"ERROR: {final_path} not found.")
        print("Run the judge first: uv run python physassistbench/run_eval.py --judge_only --language en --rule_based")
        return

    with open(final_path, encoding="utf-8") as f:
        turn_results = json.load(f)

    # Compute session correctness
    sessions: dict = defaultdict(list)
    for r in turn_results:
        sessions[r.get("test_entry_id", "")].append(r)
    session_correct_map = {
        eid: all(r.get("is_correct") for r in turns)
        for eid, turns in sessions.items()
    }

    output_path = args.output or os.path.join(results_dir, "eval_results.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    write_summary_sheet(ws_summary, turn_results, session_correct_map)
    print("  Summary sheet written")

    # Sheet 2: Examples (one per task type × subtype)
    ws_examples = wb.create_sheet("Examples")
    write_examples_sheet(ws_examples, turn_results)
    print("  Examples sheet written")

    # Sheet 3: All turns detail
    ws_detail = wb.create_sheet("All Turns")
    rows = build_rows(turn_results, session_correct_map)
    write_sheet(ws_detail, rows)
    print(f"  All Turns sheet: {len(rows)} rows")

    # Sheet 3+: One sheet per scenario
    scenarios: dict = defaultdict(list)
    for r in turn_results:
        parts = r.get("test_entry_id", "").split("_")
        sc = "_".join(parts[2:-1]) if len(parts) >= 4 else "unknown"
        scenarios[sc].append(r)

    for sc, sc_turns in sorted(scenarios.items()):
        ws = wb.create_sheet(sc[:31])
        sc_session_map = {
            eid: session_correct_map.get(eid, False)
            for eid in {r.get("test_entry_id") for r in sc_turns}
        }
        sc_rows = build_rows(sc_turns, sc_session_map)
        write_sheet(ws, sc_rows)
        print(f"  {sc}: {len(sc_rows)} rows")

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
